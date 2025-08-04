# excel/output_writer.py

import os
from openpyxl import Workbook, load_workbook
from config.settings import OUTPUT_EXCEL_PATH, OUTPUT_HEADERS
from utils.logger import log_message

def initialize_output_workbook():
    """
    Çıktı dosyası başlatılır ve başlık satırı yazılır.
    """
    if os.path.exists(OUTPUT_EXCEL_PATH):
        os.remove(OUTPUT_EXCEL_PATH)
        log_message(f"🧹 Var olan output dosyası silindi: {OUTPUT_EXCEL_PATH}")

    wb = Workbook()
    ws = wb.active
    ws.title = "btmigrate_output"
    ws.append(OUTPUT_HEADERS)
    wb.save(OUTPUT_EXCEL_PATH)
    log_message(f"✅ Output Excel dosyası oluşturuldu: {OUTPUT_EXCEL_PATH}")


def append_row_to_output(data: dict):
    """
    Belirli bir satır bilgisi Excel'e eklenir.
    """
    try:
        wb = load_workbook(OUTPUT_EXCEL_PATH)
        ws = wb.active
        row = [data.get(key, "") for key in OUTPUT_HEADERS]
        ws.append(row)
        wb.save(OUTPUT_EXCEL_PATH)
    except Exception as e:
        log_message(f"❌ Output Excel'e satır eklenemedi: {str(e)}")
